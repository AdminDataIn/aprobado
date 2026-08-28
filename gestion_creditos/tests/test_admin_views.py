from datetime import datetime, timedelta
from decimal import Decimal
import io

from django.contrib.auth import get_user_model
from django.contrib.messages import get_messages
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import connection
from django.test import TestCase
from django.test.utils import CaptureQueriesContext, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from gestion_creditos.models import (
    AsesorComercial,
    Credito,
    CreditoAdelantoNomina,
    CreditoLibranza,
    CuotaAmortizacion,
    Empresa,
    HistorialPago,
    VinculoLaboralEmpresa,
)


User = get_user_model()


class AdminViewsSmokeTest(TestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            username='staff-admin',
            email='staff-admin@aprobado.test',
            password='123456',
            is_staff=True,
        )
        self.client.login(username='staff-admin', password='123456')

    def _crear_creditos_activos_admin(self, cantidad, prefijo='CR-ADMIN-PAG'):
        usuario = User.objects.create_user(
            username=f'{prefijo.lower()}-user',
            email=f'{prefijo.lower()}@aprobado.test',
            password='123456',
        )
        for index in range(cantidad):
            Credito.objects.create(
                usuario=usuario,
                numero_credito=f'{prefijo}-{index:03d}',
                linea=Credito.LineaCredito.LIBRANZA,
                estado=Credito.EstadoCredito.ACTIVO,
                monto_solicitado=Decimal('1000000.00'),
                monto_aprobado=Decimal('1000000.00'),
                plazo_solicitado=12,
                plazo=12,
                saldo_pendiente=Decimal('500000.00'),
                capital_pendiente=Decimal('500000.00'),
                total_a_pagar=Decimal('500000.00'),
                valor_cuota=Decimal('100000.00'),
                fecha_solicitud=timezone.now() - timedelta(minutes=index),
            )

    def test_paginas_admin_principales_responden(self):
        for url_name in [
            'gestion:dashboard',
            'gestion:solicitudes',
            'gestion:adelantos_nomina',
            'gestion:creditos_activos',
            'gestion:cartera_mora',
        ]:
            response = self.client.get(reverse(url_name))
            self.assertEqual(response.status_code, 200, url_name)

    def test_detalle_credito_adelanto_renderiza_capacidad_descuento(self):
        empresa = Empresa.objects.create(
            nombre='Empresa Admin Smoke',
            tipo_empresa=Empresa.TipoEmpresa.MIXTA,
        )
        usuario = User.objects.create_user(
            username='empleado-smoke',
            email='empleado-smoke@aprobado.test',
            password='123456',
            first_name='Empleado',
            last_name='Smoke',
        )
        vinculo = VinculoLaboralEmpresa.objects.create(
            usuario=usuario,
            empresa=empresa,
            documento_empleado='123456789',
            nombre_empleado='EMPLEADO SMOKE',
            estado_vinculo=VinculoLaboralEmpresa.EstadoVinculo.ACTIVO,
            fecha_alta_aprobado=timezone.localdate() - timedelta(days=90),
            salario_base_mensual=Decimal('2400000.00'),
            auxilio_transporte_mensual=Decimal('162000.00'),
            descuentos_fijos_mensuales=Decimal('350000.00'),
        )
        credito = Credito.objects.create(
            usuario=usuario,
            numero_credito='CR-ADMIN-SMOKE-0001',
            linea=Credito.LineaCredito.ADELANTO_NOMINA,
            estado=Credito.EstadoCredito.SOLICITUD,
            monto_solicitado=Decimal('500000.00'),
            plazo_solicitado=1,
        )
        CreditoAdelantoNomina.objects.create(
            credito=credito,
            vinculo_laboral=vinculo,
            monto_solicitado=Decimal('500000.00'),
            monto_maximo_calculado=Decimal('600000.00'),
            dias_adelanto=5,
            salario_base_usado=Decimal('2400000.00'),
        )

        response = self.client.get(reverse('gestion:credito_detalle', args=[credito.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Capacidad de descuento')

    def test_dashboard_export_descarga_excel_funcional(self):
        response = self.client.get(reverse('gestion:dashboard_export'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        workbook = load_workbook(io.BytesIO(response.content))
        self.assertIn('Resumen', workbook.sheetnames)
        self.assertIn('Solicitudes', workbook.sheetnames)
        self.assertIn('Creditos', workbook.sheetnames)
        self.assertIn('Cuotas', workbook.sheetnames)
        self.assertIn('Pagos', workbook.sheetnames)
        self.assertIn('Recaudo contable', workbook.sheetnames)
        self.assertIn('Detalle contable', workbook.sheetnames)

    def test_dashboard_renderiza_indicadores_contables(self):
        response = self.client.get(reverse('gestion:dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Total recaudado')
        self.assertContains(response, 'Capital recuperado')

    def test_admin_creditos_activos_responde_pagina_tres(self):
        self._crear_creditos_activos_admin(45)

        response = self.client.get(reverse('gestion:creditos_activos'), {'page': 3})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['creditos'].number, 3)
        self.assertEqual(response.context['creditos'].paginator.num_pages, 3)

    def test_admin_creditos_activos_no_duplica_page_y_preserva_filtros(self):
        self._crear_creditos_activos_admin(45, prefijo='CR-ADMIN-FILTRO')

        response = self.client.get(
            reverse('gestion:creditos_activos'),
            {'page': 2, 'search': 'CR-ADMIN-FILTRO', 'linea': Credito.LineaCredito.LIBRANZA},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'page=3&search=CR-ADMIN-FILTRO&amp;linea=LIBRANZA', html=False)
        self.assertNotContains(response, 'page=3&amp;page=2', html=False)

    def test_admin_creditos_activos_url_malformada_usa_primer_page(self):
        self._crear_creditos_activos_admin(45, prefijo='CR-ADMIN-DUP')

        response = self.client.get(
            f"{reverse('gestion:creditos_activos')}?page=3&page=2&search=CR-ADMIN-DUP"
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['creditos'].number, 3)
        self.assertNotContains(response, 'page=2&amp;page=3', html=False)

    @override_settings(MANUAL_PAYMENT_AUTH_KEY='clave-prueba-admin')
    def test_pago_manual_admin_total_cierra_credito_activo(self):
        credito = Credito.objects.create(
            usuario=self.staff,
            numero_credito='CR-ADMIN-PAGO-TOTAL',
            linea=Credito.LineaCredito.LIBRANZA,
            estado=Credito.EstadoCredito.ACTIVO,
            monto_solicitado=Decimal('3000000.00'),
            monto_aprobado=Decimal('3000000.00'),
            plazo_solicitado=1,
            plazo=1,
            saldo_pendiente=Decimal('3357000.00'),
            capital_pendiente=Decimal('3000000.00'),
            comision=Decimal('300000.00'),
            iva_comision=Decimal('57000.00'),
            tasa_interes=Decimal('1.90'),
            valor_cuota=Decimal('3420783.00'),
            total_a_pagar=Decimal('3420783.00'),
            fecha_proximo_pago=timezone.localdate(),
        )
        cuota = CuotaAmortizacion.objects.create(
            credito=credito,
            numero_cuota=1,
            fecha_vencimiento=timezone.localdate(),
            capital_a_pagar=Decimal('3357000.00'),
            interes_a_pagar=Decimal('63783.00'),
            valor_cuota=Decimal('3420783.00'),
            saldo_capital_pendiente=Decimal('0.00'),
        )

        response = self.client.post(
            reverse('gestion:credito_agregar_pago', args=[credito.id]),
            {
                'monto': '3420783.00',
                'referencia_pago': 'ADMIN-PAGO-TOTAL-001',
                'metodo_pago': HistorialPago.MetodoPago.OFFLINE_MANUAL,
                'auth_key': 'clave-prueba-admin',
            },
        )

        self.assertEqual(response.status_code, 302)
        credito.refresh_from_db()
        cuota.refresh_from_db()
        self.assertEqual(credito.estado, Credito.EstadoCredito.PAGADO)
        self.assertTrue(cuota.pagada)
        self.assertEqual(HistorialPago.objects.filter(credito=credito).count(), 1)
        mensajes = [str(mensaje) for mensaje in get_messages(response.wsgi_request)]
        self.assertFalse(any('error inesperado' in mensaje.lower() for mensaje in mensajes))

    @override_settings(MANUAL_PAYMENT_AUTH_KEY='clave-prueba-admin')
    def test_pago_manual_admin_bloquea_credito_ya_pagado_con_mensaje_controlado(self):
        credito = Credito.objects.create(
            usuario=self.staff,
            numero_credito='CR-ADMIN-YA-PAGADO',
            linea=Credito.LineaCredito.LIBRANZA,
            estado=Credito.EstadoCredito.PAGADO,
            monto_solicitado=Decimal('100000.00'),
            monto_aprobado=Decimal('100000.00'),
            plazo_solicitado=1,
            plazo=1,
            saldo_pendiente=Decimal('0.00'),
            capital_pendiente=Decimal('0.00'),
        )

        response = self.client.post(
            reverse('gestion:credito_agregar_pago', args=[credito.id]),
            {
                'monto': '100000.00',
                'auth_key': 'clave-prueba-admin',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(HistorialPago.objects.filter(credito=credito).exists())
        mensajes = [str(mensaje) for mensaje in get_messages(response.wsgi_request)]
        self.assertIn(
            'El crédito ya se encuentra pagado y no permite registrar pagos manuales adicionales.',
            mensajes,
        )

    def test_historial_pago_legacy_sin_comprobante_muestra_estado_claro(self):
        credito = Credito.objects.create(
            usuario=self.staff,
            numero_credito='CR-LEG-SIN-SOP',
            linea=Credito.LineaCredito.LIBRANZA,
            estado=Credito.EstadoCredito.ACTIVO,
            monto_solicitado=Decimal('730000.00'),
            monto_aprobado=Decimal('730000.00'),
            plazo_solicitado=1,
            plazo=1,
            saldo_pendiente=Decimal('66.21'),
            capital_pendiente=Decimal('66.21'),
        )
        HistorialPago.objects.create(
            credito=credito,
            monto=Decimal('730000.00'),
            referencia_pago='MANUAL-LEGACY-SIN-SOPORTE',
            estado=HistorialPago.EstadoPago.EXITOSO,
            metodo_pago=HistorialPago.MetodoPago.NO_DEFINIDO,
            origen_registro=HistorialPago.OrigenRegistro.LEGACY,
        )

        response = self.client.get(reverse('gestion:credito_detalle', args=[credito.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Sin comprobante registrado')
        self.assertContains(response, 'Adjuntar comprobante')

    def test_staff_adjunta_comprobante_a_pago_existente_sin_crear_otro_pago(self):
        credito = Credito.objects.create(
            usuario=self.staff,
            numero_credito='CR-LEG-SOPORTE',
            linea=Credito.LineaCredito.LIBRANZA,
            estado=Credito.EstadoCredito.ACTIVO,
            monto_solicitado=Decimal('730000.00'),
            monto_aprobado=Decimal('730000.00'),
            plazo_solicitado=1,
            plazo=1,
            saldo_pendiente=Decimal('66.21'),
            capital_pendiente=Decimal('66.21'),
        )
        pago = HistorialPago.objects.create(
            credito=credito,
            monto=Decimal('730000.00'),
            referencia_pago='MANUAL-LEGACY-SOPORTE',
            estado=HistorialPago.EstadoPago.EXITOSO,
            metodo_pago=HistorialPago.MetodoPago.NO_DEFINIDO,
            origen_registro=HistorialPago.OrigenRegistro.LEGACY,
        )
        monto_original = pago.monto

        response = self.client.post(
            reverse('gestion:pago_comprobante_actualizar', args=[pago.id]),
            {
                'comprobante': SimpleUploadedFile(
                    'comprobante-legacy.pdf',
                    b'%PDF-1.4 comprobante de prueba',
                    content_type='application/pdf',
                ),
            },
        )

        self.assertEqual(response.status_code, 302)
        pago.refresh_from_db()
        self.assertTrue(pago.comprobante)
        self.assertIn('comprobante-legacy', pago.comprobante.name)
        self.assertEqual(pago.monto, monto_original)
        self.assertEqual(HistorialPago.objects.filter(credito=credito).count(), 1)

        response_archivo = self.client.get(reverse('gestion:pago_comprobante', args=[pago.id]))
        self.assertEqual(response_archivo.status_code, 200)
        self.assertEqual(response_archivo['Content-Type'], 'application/pdf')

    def test_usuario_no_staff_no_puede_actualizar_comprobante_legacy(self):
        usuario = User.objects.create_user(
            username='usuario-sin-permiso-comprobante',
            email='sin-permiso@aprobado.test',
            password='123456',
        )
        credito = Credito.objects.create(
            usuario=usuario,
            numero_credito='CR-SOP-PROTEGIDO',
            linea=Credito.LineaCredito.LIBRANZA,
            estado=Credito.EstadoCredito.ACTIVO,
            monto_solicitado=Decimal('100000.00'),
            monto_aprobado=Decimal('100000.00'),
            plazo_solicitado=1,
            plazo=1,
        )
        pago = HistorialPago.objects.create(
            credito=credito,
            monto=Decimal('100000.00'),
            referencia_pago='MANUAL-SOPORTE-PROTEGIDO',
            estado=HistorialPago.EstadoPago.EXITOSO,
        )
        self.client.force_login(usuario)

        response = self.client.post(
            reverse('gestion:pago_comprobante_actualizar', args=[pago.id]),
            {
                'comprobante': SimpleUploadedFile(
                    'no-autorizado.pdf',
                    b'%PDF-1.4 no autorizado',
                    content_type='application/pdf',
                ),
            },
        )

        self.assertEqual(response.status_code, 302)
        pago.refresh_from_db()
        self.assertFalse(pago.comprobante)


class AdminCreditosActivosFiltersTest(TestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            username='staff-filtros-activos',
            password='123456',
            is_staff=True,
        )
        self.usuario = User.objects.create_user(
            username='cliente-filtros-activos',
            email='cliente-filtros@aprobado.test',
            password='123456',
            first_name='Cliente',
            last_name='Filtros',
        )
        self.asesor = AsesorComercial.objects.create(
            nombre='Asesor Filtros',
            cedula='900001234',
            activo=True,
        )
        self.empresa_a = Empresa.objects.create(
            nombre='Empresa Filtros A',
            asesor_comercial=self.asesor,
        )
        self.empresa_b = Empresa.objects.create(nombre='Empresa Filtros B')
        self.client.force_login(self.staff)

    def _crear_credito(
        self,
        numero,
        *,
        estado=Credito.EstadoCredito.ACTIVO,
        linea=Credito.LineaCredito.LIBRANZA,
        empresa=None,
        fecha_desembolso=None,
    ):
        credito = Credito.objects.create(
            usuario=self.usuario,
            numero_credito=numero,
            linea=linea,
            estado=estado,
            monto_solicitado=Decimal('1000000.00'),
            monto_aprobado=Decimal('1000000.00'),
            plazo_solicitado=6,
            plazo=6,
            saldo_pendiente=Decimal('800000.00'),
            capital_pendiente=Decimal('750000.00'),
            total_a_pagar=Decimal('1100000.00'),
            valor_cuota=Decimal('190000.00'),
            fecha_desembolso=fecha_desembolso,
        )
        if linea == Credito.LineaCredito.LIBRANZA:
            CreditoLibranza.objects.create(
                credito=credito,
                empresa=empresa or self.empresa_a,
                direccion='Calle 10',
                telefono='3001234567',
                correo_electronico='cliente@empresa.test',
                cedula='1000000000',
                nombres='Cliente',
                apellidos='Filtros',
            )
        return credito

    def _fecha(self, year, month, day):
        return timezone.make_aware(datetime(year, month, day, 9, 0))

    def _crear_universo(self):
        activo_a = self._crear_credito(
            'CR-FLT-A-001',
            empresa=self.empresa_a,
            fecha_desembolso=self._fecha(2026, 8, 10),
        )
        activo_b = self._crear_credito(
            'CR-FLT-B-001',
            empresa=self.empresa_b,
            fecha_desembolso=self._fecha(2026, 7, 10),
        )
        mora_a = self._crear_credito(
            'CR-FLT-M-001',
            empresa=self.empresa_a,
            estado=Credito.EstadoCredito.EN_MORA,
            fecha_desembolso=self._fecha(2026, 8, 12),
        )
        emprendimiento = self._crear_credito(
            'CR-FLT-E-001',
            linea=Credito.LineaCredito.EMPRENDIMIENTO,
            fecha_desembolso=self._fecha(2026, 8, 15),
        )
        return activo_a, activo_b, mora_a, emprendimiento

    def test_vista_requiere_staff(self):
        url = reverse('gestion:creditos_activos')
        self.client.logout()
        self.assertEqual(self.client.get(url).status_code, 302)
        usuario_normal = User.objects.create_user(
            username='usuario-filtros-activos',
            password='123456',
        )
        self.client.force_login(usuario_normal)
        self.assertEqual(self.client.get(url).status_code, 302)

    def test_filtros_avanzados_preservan_valores_y_usan_fecha_desembolso(self):
        activo_a, _activo_b, _mora_a, _emprendimiento = self._crear_universo()
        response = self.client.get(reverse('gestion:creditos_activos'), {
            'search': activo_a.numero_credito,
            'fecha_desde': '2026-08-01',
            'fecha_hasta': '2026-08-31',
            'empresa': self.empresa_a.nombre,
            'estado': Credito.EstadoCredito.ACTIVO,
            'linea': Credito.LineaCredito.LIBRANZA,
            'asesor': str(self.asesor.pk),
        })

        self.assertEqual(response.context['creditos'].paginator.count, 1)
        self.assertEqual(response.context['creditos'][0].pk, activo_a.pk)
        self.assertContains(response, 'aria-expanded="true"', html=False)
        self.assertContains(response, 'value="2026-08-01"', html=False)
        self.assertContains(response, 'Fecha desembolso')
        self.assertContains(response, 'Limpiar filtros')

    def test_empresa_estado_linea_asesor_y_fechas_filtran_en_queryset(self):
        activo_a, activo_b, mora_a, emprendimiento = self._crear_universo()
        casos = (
            ({'empresa': self.empresa_a.nombre}, {activo_a.pk}),
            ({'estado': Credito.EstadoCredito.EN_MORA}, {mora_a.pk}),
            ({'linea': Credito.LineaCredito.EMPRENDIMIENTO}, {emprendimiento.pk}),
            ({'asesor': str(self.asesor.pk)}, {activo_a.pk}),
            (
                {'fecha_desde': '2026-07-01', 'fecha_hasta': '2026-07-31'},
                {activo_b.pk},
            ),
        )
        for params, esperados in casos:
            with self.subTest(params=params):
                response = self.client.get(reverse('gestion:creditos_activos'), params)
                encontrados = {credito.pk for credito in response.context['creditos']}
                self.assertEqual(encontrados, esperados)

    def test_rango_invalido_muestra_error_sin_aplicar_fechas(self):
        activo_a, activo_b, _mora_a, emprendimiento = self._crear_universo()
        response = self.client.get(reverse('gestion:creditos_activos'), {
            'fecha_desde': '2026-09-01',
            'fecha_hasta': '2026-08-01',
        })

        self.assertTrue(response.context['filtros_errores'])
        self.assertEqual(
            {credito.pk for credito in response.context['creditos']},
            {activo_a.pk, activo_b.pk, emprendimiento.pk},
        )
        self.assertContains(response, 'La fecha inicial no puede ser posterior')

    def test_sin_get_mantiene_activo_por_defecto_y_filtros_cerrados(self):
        activo_a, activo_b, _mora_a, emprendimiento = self._crear_universo()
        response = self.client.get(reverse('gestion:creditos_activos'))

        self.assertEqual(
            {credito.pk for credito in response.context['creditos']},
            {activo_a.pk, activo_b.pk, emprendimiento.pk},
        )
        self.assertContains(response, 'aria-expanded="false"', html=False)
        self.assertContains(response, reverse('gestion:creditos_activos'))

    def test_consultas_no_crecen_por_credito_renderizado(self):
        self._crear_credito(
            'CR-N1-001',
            fecha_desembolso=self._fecha(2026, 8, 10),
        )
        url = reverse('gestion:creditos_activos')
        with CaptureQueriesContext(connection) as consultas_base:
            self.client.get(url)

        for indice in range(2, 10):
            self._crear_credito(
                f'CR-N1-{indice:03d}',
                fecha_desembolso=self._fecha(2026, 8, 10),
            )
        with CaptureQueriesContext(connection) as consultas_ampliadas:
            self.client.get(url)

        self.assertLessEqual(len(consultas_ampliadas), len(consultas_base) + 1)
