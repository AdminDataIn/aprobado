import io
from datetime import datetime, timedelta
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.db import connection
from django.test import RequestFactory, TestCase
from django.test.utils import CaptureQueriesContext
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from gestion_creditos.models import (
    AsesorComercial,
    Credito,
    CreditoLibranza,
    CuotaAmortizacion,
    DetalleContablePago,
    Empresa,
    HistorialPago,
)
from gestion_creditos.services.admin_excel_report import build_admin_excel_report


User = get_user_model()


class AdminExcelReportTests(TestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.staff = User.objects.create_user(
            username='excel-admin',
            password='123456',
            is_staff=True,
        )
        self.cliente = User.objects.create_user(
            username='excel-cliente',
            password='123456',
            first_name='Cliente',
            last_name='Reporte',
        )
        self.asesor_a = AsesorComercial.objects.create(
            nombre='Ejecutivo Excel A',
            cedula='910000001',
        )
        self.asesor_b = AsesorComercial.objects.create(
            nombre='Ejecutivo Excel B',
            cedula='910000002',
        )
        self.empresa_a = Empresa.objects.create(
            nombre='Empresa Excel A',
            asesor_comercial=self.asesor_a,
        )
        self.empresa_b = Empresa.objects.create(
            nombre='Empresa Excel B',
            asesor_comercial=self.asesor_b,
        )
        self.client.force_login(self.staff)
        self._sequence = 0

    def _aware(self, year, month, day):
        return timezone.make_aware(datetime(year, month, day, 10, 0))

    def _crear_credito(
        self,
        *,
        empresa=None,
        estado=Credito.EstadoCredito.SOLICITUD,
        linea=Credito.LineaCredito.LIBRANZA,
        fecha_solicitud=None,
        fecha_desembolso=None,
        usuario=None,
    ):
        self._sequence += 1
        credito = Credito.objects.create(
            usuario=usuario or self.cliente,
            numero_credito=f'CR-EX-{self._sequence:04d}',
            linea=linea,
            estado=estado,
            monto_solicitado=Decimal('1000000.00'),
            monto_aprobado=Decimal('900000.00'),
            plazo_solicitado=6,
            plazo=6,
            tasa_interes=Decimal('2.20'),
            saldo_pendiente=Decimal('750000.00'),
            capital_pendiente=Decimal('700000.00'),
            valor_cuota=Decimal('170000.00'),
            total_a_pagar=Decimal('1020000.00'),
            fecha_desembolso=fecha_desembolso,
            fecha_proximo_pago=timezone.localdate() + timedelta(days=10),
        )
        if fecha_solicitud:
            Credito.objects.filter(pk=credito.pk).update(fecha_solicitud=fecha_solicitud)
            credito.fecha_solicitud = fecha_solicitud
        if linea == Credito.LineaCredito.LIBRANZA:
            CreditoLibranza.objects.create(
                credito=credito,
                empresa=empresa or self.empresa_a,
                nombres=(usuario or self.cliente).first_name or 'Cliente',
                apellidos=(usuario or self.cliente).last_name or 'Reporte',
                cedula='1000123456',
                direccion='Calle privada',
                telefono='3001234567',
                correo_electronico='cliente@reporte.test',
            )
        return credito

    def _crear_cuota(self, credito, vencimiento, numero=1):
        return CuotaAmortizacion.objects.create(
            credito=credito,
            numero_cuota=numero,
            fecha_vencimiento=vencimiento,
            capital_a_pagar=Decimal('150000.00'),
            interes_a_pagar=Decimal('20000.00'),
            valor_cuota=Decimal('170000.00'),
            saldo_capital_pendiente=Decimal('600000.00'),
            monto_pagado=Decimal('20000.00'),
        )

    def _crear_pago(self, credito, cuota, fecha, referencia):
        pago = HistorialPago.objects.create(
            credito=credito,
            fecha_aplicacion=fecha,
            monto=Decimal('170000.00'),
            referencia_pago=referencia,
            estado=HistorialPago.EstadoPago.EXITOSO,
            metodo_pago=HistorialPago.MetodoPago.OFFLINE_MANUAL,
            origen_registro=HistorialPago.OrigenRegistro.REGISTRO_MANUAL_ADMIN,
        )
        DetalleContablePago.objects.create(
            pago=pago,
            credito=credito,
            cuota=cuota,
            fecha_aplicacion=fecha,
            monto_total_aplicado=Decimal('170000.00'),
            capital_aplicado=Decimal('150000.00'),
            capital_principal_aplicado=Decimal('140000.00'),
            interes_aplicado=Decimal('20000.00'),
            comision_aplicada=Decimal('8000.00'),
            iva_aplicado=Decimal('2000.00'),
        )
        return pago

    def _download(self, params=None):
        response = self.client.get(reverse('gestion:dashboard_export'), params or {})
        workbook = None
        if response.status_code == 200:
            workbook = load_workbook(io.BytesIO(response.content))
        return response, workbook

    def _sheet_rows(self, sheet):
        headers = [cell.value for cell in sheet[1]]
        return [dict(zip(headers, values)) for values in sheet.iter_rows(min_row=2, values_only=True)]

    def test_exportar_sin_filtros_incluye_hojas_y_columnas_auditables(self):
        self._crear_credito()

        response, workbook = self._download()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            workbook.sheetnames,
            [
                'Resumen',
                'Solicitudes',
                'Creditos',
                'Cuotas',
                'Pagos',
                'Cartera por linea',
                'Creditos por estado',
                'Distribucion empresas',
                'Recaudo contable',
                'Detalle contable',
            ],
        )
        self.assertIn('Documento enmascarado', [cell.value for cell in workbook['Solicitudes'][1]])
        self.assertIn('Capital pendiente actual', [cell.value for cell in workbook['Creditos'][1]])
        self.assertEqual(workbook['Solicitudes'].freeze_panes, 'A2')
        self.assertTrue(workbook['Solicitudes'].auto_filter.ref)
        solicitudes_headers = {
            cell.value: cell.column for cell in workbook['Solicitudes'][1]
        }
        self.assertEqual(
            workbook['Solicitudes'].cell(
                row=2,
                column=solicitudes_headers['Monto solicitado'],
            ).number_format,
            '$#,##0.00',
        )
        self.assertIn(
            'dd/mm/yyyy',
            workbook['Solicitudes'].cell(
                row=2,
                column=solicitudes_headers['Fecha solicitud'],
            ).number_format,
        )

    def test_cada_hoja_aplica_la_fecha_de_su_dominio(self):
        solicitud = self._crear_credito(fecha_solicitud=self._aware(2026, 8, 5))
        desembolsado = self._crear_credito(
            estado=Credito.EstadoCredito.ACTIVO,
            fecha_solicitud=self._aware(2026, 7, 20),
            fecha_desembolso=self._aware(2026, 8, 10),
        )
        cuota_agosto = self._crear_cuota(desembolsado, datetime(2026, 8, 25).date())
        self._crear_cuota(desembolsado, datetime(2026, 9, 25).date(), numero=2)
        self._crear_pago(desembolsado, cuota_agosto, self._aware(2026, 8, 20), 'REF-EX-IN')
        self._crear_pago(desembolsado, cuota_agosto, self._aware(2026, 9, 20), 'REF-EX-OUT')

        _response, workbook = self._download({
            'fecha_desde': '2026-08-01',
            'fecha_hasta': '2026-08-31',
        })

        solicitudes = self._sheet_rows(workbook['Solicitudes'])
        creditos = self._sheet_rows(workbook['Creditos'])
        cuotas = self._sheet_rows(workbook['Cuotas'])
        pagos = self._sheet_rows(workbook['Pagos'])
        self.assertEqual([row['Número solicitud / crédito'] for row in solicitudes], [solicitud.numero_credito])
        self.assertEqual([row['Número crédito'] for row in creditos], [desembolsado.numero_credito])
        self.assertEqual([row['Número cuota'] for row in cuotas], [1])
        self.assertEqual([row['Referencia / transacción'] for row in pagos], ['REF-EX-IN'])

    def test_filtros_empresa_estado_linea_asesor_y_combinacion(self):
        target = self._crear_credito(
            empresa=self.empresa_a,
            estado=Credito.EstadoCredito.ACTIVO,
            fecha_solicitud=self._aware(2026, 8, 5),
        )
        otro_empresa = self._crear_credito(
            empresa=self.empresa_b,
            estado=Credito.EstadoCredito.ACTIVO,
            fecha_solicitud=self._aware(2026, 8, 6),
        )
        rechazado = self._crear_credito(
            empresa=self.empresa_a,
            estado=Credito.EstadoCredito.RECHAZADO,
            fecha_solicitud=self._aware(2026, 8, 7),
        )
        emprendimiento = self._crear_credito(
            linea=Credito.LineaCredito.EMPRENDIMIENTO,
            estado=Credito.EstadoCredito.ACTIVO,
            fecha_solicitud=self._aware(2026, 8, 8),
        )
        casos = (
            ({'empresa': str(self.empresa_b.pk)}, {otro_empresa.numero_credito}),
            ({'estado': Credito.EstadoCredito.RECHAZADO}, {rechazado.numero_credito}),
            ({'linea': Credito.LineaCredito.EMPRENDIMIENTO}, {emprendimiento.numero_credito}),
            ({'asesor': str(self.asesor_a.pk)}, {target.numero_credito, rechazado.numero_credito}),
            ({
                'empresa': str(self.empresa_a.pk),
                'estado': Credito.EstadoCredito.ACTIVO,
                'linea': Credito.LineaCredito.LIBRANZA,
                'asesor': str(self.asesor_a.pk),
            }, {target.numero_credito}),
        )

        for params, expected in casos:
            with self.subTest(params=params):
                _response, workbook = self._download(params)
                actual = {
                    row['Número solicitud / crédito']
                    for row in self._sheet_rows(workbook['Solicitudes'])
                }
                self.assertEqual(actual, expected)

    def test_rango_invalido_no_genera_archivo(self):
        response = self.client.get(reverse('gestion:dashboard_export'), {
            'fecha_desde': '2026-09-01',
            'fecha_hasta': '2026-08-01',
        })

        self.assertEqual(response.status_code, 400)
        self.assertNotEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertContains(response, 'fecha inicial', status_code=400)

    def test_resumen_documenta_filtros_y_saldos_como_corte_actual(self):
        self._crear_credito(
            empresa=self.empresa_a,
            estado=Credito.EstadoCredito.ACTIVO,
            fecha_solicitud=self._aware(2026, 8, 5),
        )

        _response, workbook = self._download({
            'fecha_desde': '2026-08-01',
            'fecha_hasta': '2026-08-31',
            'empresa': str(self.empresa_a.pk),
            'estado': Credito.EstadoCredito.ACTIVO,
            'linea': Credito.LineaCredito.LIBRANZA,
            'asesor': str(self.asesor_a.pk),
        })

        resumen = {row[0]: row[1] for row in workbook['Resumen'].iter_rows(min_row=2, values_only=True)}
        self.assertEqual(resumen['Empresa'], self.empresa_a.nombre)
        self.assertEqual(resumen['Estado'], Credito.EstadoCredito.ACTIVO)
        self.assertEqual(resumen['Línea / producto'], Credito.LineaCredito.LIBRANZA)
        self.assertEqual(resumen['Ejecutivo / asesor'], self.asesor_a.nombre)
        self.assertIn('Saldo total cartera - corte actual', resumen)
        self.assertIn('Saldo capital pendiente - corte actual', resumen)
        self.assertIn('fecha de desembolso', resumen['Semántica créditos'])

    def test_strings_de_usuario_no_se_convierten_en_formulas(self):
        usuario = User.objects.create_user(
            username='excel-formula',
            password='123456',
            first_name='=SUM(1,1)',
            last_name='Prueba',
        )
        credito = self._crear_credito(usuario=usuario)

        _response, workbook = self._download()

        row = next(
            item for item in self._sheet_rows(workbook['Solicitudes'])
            if item['Número solicitud / crédito'] == credito.numero_credito
        )
        self.assertTrue(row['Solicitante'].startswith("'="))
        self.assertNotIn('1000123456', row['Documento enmascarado'])
        self.assertTrue(row['Documento enmascarado'].endswith('3456'))

    def test_dashboard_exportacion_usa_modal_independiente(self):
        self._crear_credito(empresa=self.empresa_a)
        response = self.client.get(reverse('gestion:dashboard'), {
            'fecha_desde': '2026-08-01',
            'fecha_hasta': '2026-08-31',
            'empresa': self.empresa_a.nombre,
            'estado': Credito.EstadoCredito.ACTIVO,
            'linea': Credito.LineaCredito.LIBRANZA,
            'asesor': str(self.asesor_a.pk),
        })

        export_url = reverse('gestion:dashboard_export')
        self.assertContains(response, 'id="openExportReportModal"', html=False)
        self.assertContains(response, 'data-bs-target="#exportReportModal"', html=False)
        self.assertNotContains(response, f'href="{export_url}', html=False)
        self.assertContains(response, 'id="exportReportModal"', html=False)
        self.assertContains(response, f'action="{export_url}"', html=False)
        self.assertContains(response, 'data-loader="off"', html=False)
        self.assertContains(response, 'value="general" checked', html=False)
        self.assertContains(response, 'value="range"', html=False)
        self.assertContains(response, 'name="fecha_desde" type="date"', html=False)
        self.assertContains(response, 'name="fecha_hasta" type="date"', html=False)
        self.assertContains(response, 'exportFechaDesde.value > exportFechaHasta.value', html=False)
        self.assertContains(response, 'Configura la descarga sin modificar los filtros del dashboard.', html=False)
        self.assertContains(response, f'value="{self.empresa_a.nombre}" selected', html=False)
        self.assertContains(response, 'value="ACTIVO" selected', html=False)

    def test_exportacion_requiere_staff(self):
        url = reverse('gestion:dashboard_export')
        self.client.logout()
        self.assertEqual(self.client.get(url).status_code, 302)
        normal = User.objects.create_user(username='excel-normal', password='123456')
        self.client.force_login(normal)
        self.assertEqual(self.client.get(url).status_code, 302)

    def test_consultas_no_crecen_linealmente_por_credito(self):
        self._crear_credito(fecha_solicitud=self._aware(2026, 8, 1))
        request = self.factory.get('/gestion/exportar-reporte/')
        with CaptureQueriesContext(connection) as base_queries:
            build_admin_excel_report(request)

        for day in range(2, 8):
            self._crear_credito(fecha_solicitud=self._aware(2026, 8, day))
        with CaptureQueriesContext(connection) as expanded_queries:
            build_admin_excel_report(request)

        self.assertLessEqual(len(expanded_queries), len(base_queries) + 2)
