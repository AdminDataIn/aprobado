from decimal import Decimal

from django.db.models import Count, DecimalField, OuterRef, Subquery, Sum, Value
from django.db.models.functions import Coalesce
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from gestion_creditos.models import (
    Credito,
    CuotaAmortizacion,
    DetalleContablePago,
    HistorialEstado,
    HistorialPago,
)
from gestion_creditos.services.admin_dashboard_filters import parse_admin_dashboard_filters
from gestion_creditos.services.dashboard_metrics import ESTADOS_CARTERA, _base_admin_queryset


EXCEL_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
MONEY_FORMAT = '$#,##0.00'
PERCENT_FORMAT = '0.00%'
DATE_FORMAT = 'dd/mm/yyyy'
DATETIME_FORMAT = 'dd/mm/yyyy hh:mm'
HEADER_FILL = PatternFill('solid', fgColor='0F2742')
HEADER_FONT = Font(color='FFFFFF', bold=True)
CURRENCY_HEADERS = {
    'Monto solicitado',
    'Monto aprobado',
    'Monto desembolsado',
    'Valor cuota',
    'Monto pagado',
    'Saldo pendiente cuota',
    'Valor total',
    'Capital',
    'Interés',
    'Comisión',
    'IVA',
    'Saldo total actual',
    'Capital pendiente actual',
    'Total recaudado',
    'Capital recuperado',
    'Interés recuperado',
    'Comisión recuperada',
    'IVA recuperado',
}
DATE_HEADERS = {
    'Fecha solicitud',
    'Fecha aprobación',
    'Fecha rechazo',
    'Fecha desembolso',
    'Próxima fecha pago',
    'Vencimiento',
    'Fecha aplicación',
}


class AdminExcelReportError(ValueError):
    pass


def build_admin_excel_report(request):
    filtros = parse_admin_dashboard_filters(request)
    if filtros.errores:
        raise AdminExcelReportError(' '.join(filtros.errores))

    base_creditos = filtros.aplicar_dimensiones_credito(_base_admin_queryset()).select_related(
        'detalle_libranza__empresa__asesor_comercial',
        'detalle_adelanto_nomina__vinculo_laboral__empresa__asesor_comercial',
    )
    base_creditos = _annotate_decision_dates(base_creditos)
    solicitudes = filtros.aplicar_fecha_credito(base_creditos, 'fecha_solicitud').order_by(
        'fecha_solicitud', 'pk'
    )
    desembolsos = filtros.aplicar_fecha_credito(
        base_creditos.filter(fecha_desembolso__isnull=False),
        'fecha_desembolso',
    ).order_by('fecha_desembolso', 'pk')

    cuotas = filtros.aplicar_fecha_vencimiento(
        CuotaAmortizacion.objects.filter(credito__in=base_creditos).select_related(
            'credito__usuario',
            'credito__detalle_libranza__empresa__asesor_comercial',
            'credito__detalle_emprendimiento',
            'credito__detalle_adelanto_nomina__vinculo_laboral__empresa__asesor_comercial',
        )
    ).order_by('fecha_vencimiento', 'credito__numero_credito', 'numero_cuota')

    pagos = filtros.aplicar_fecha_recaudo(
        HistorialPago.objects.filter(
            credito__in=base_creditos,
            estado=HistorialPago.EstadoPago.EXITOSO,
        ).select_related(
            'credito__usuario',
            'credito__detalle_libranza__empresa__asesor_comercial',
            'credito__detalle_emprendimiento',
            'credito__detalle_adelanto_nomina__vinculo_laboral__empresa__asesor_comercial',
        ).annotate(
            capital_reporte=Coalesce(
                Sum('detalles_contables__capital_principal_aplicado'),
                _money_zero(),
            ),
            interes_reporte=Coalesce(
                Sum('detalles_contables__interes_aplicado'),
                _money_zero(),
            ),
            comision_reporte=Coalesce(
                Sum('detalles_contables__comision_aplicada'),
                _money_zero(),
            ),
            iva_reporte=Coalesce(
                Sum('detalles_contables__iva_aplicado'),
                _money_zero(),
            ),
        )
    ).order_by('fecha_aplicacion', 'pk')

    detalles = filtros.aplicar_fecha_recaudo(
        DetalleContablePago.objects.filter(credito__in=base_creditos).select_related(
            'credito',
            'credito__detalle_libranza__empresa',
            'credito__detalle_adelanto_nomina__vinculo_laboral__empresa',
            'pago',
            'cuota',
        )
    ).order_by('credito__numero_credito', 'fecha_aplicacion', 'secuencia_aplicacion')

    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_resumen(
        workbook,
        filtros=filtros,
        base_creditos=base_creditos,
        solicitudes=solicitudes,
        desembolsos=desembolsos,
        pagos=pagos,
        detalles=detalles,
    )
    _write_solicitudes(workbook, solicitudes)
    _write_creditos(workbook, desembolsos)
    _write_cuotas(workbook, cuotas)
    _write_pagos(workbook, pagos)
    _write_secondary_sheets(workbook, base_creditos, detalles)

    suffix = _report_range_suffix(filtros)
    return workbook, f'reporte_administrativo_{suffix}.xlsx'


def _annotate_decision_dates(queryset):
    aprobacion = (
        HistorialEstado.objects.filter(
            credito_id=OuterRef('pk'),
            estado_nuevo__in=[
                Credito.EstadoCredito.APROBADO_PAGADOR,
                Credito.EstadoCredito.APROBADO,
            ],
        )
        .order_by('fecha', 'pk')
        .values('fecha')[:1]
    )
    rechazo = (
        HistorialEstado.objects.filter(
            credito_id=OuterRef('pk'),
            estado_nuevo=Credito.EstadoCredito.RECHAZADO,
        )
        .order_by('fecha', 'pk')
        .values('fecha')[:1]
    )
    return queryset.annotate(
        fecha_aprobacion_reporte=Subquery(aprobacion),
        fecha_rechazo_reporte=Subquery(rechazo),
    )


def _write_resumen(
    workbook,
    *,
    filtros,
    base_creditos,
    solicitudes,
    desembolsos,
    pagos,
    detalles,
):
    cartera_actual = base_creditos.filter(estado__in=ESTADOS_CARTERA).aggregate(
        saldo=Coalesce(Sum('saldo_pendiente'), _money_zero()),
        capital=Coalesce(Sum('capital_pendiente'), _money_zero()),
    )
    desembolsos_totales = desembolsos.aggregate(
        monto=Coalesce(Sum('monto_aprobado'), _money_zero()),
    )
    pagos_totales = pagos.aggregate(
        monto=Coalesce(Sum('monto'), _money_zero()),
    )
    recaudo = detalles.aggregate(
        total=Coalesce(Sum('monto_total_aplicado'), _money_zero()),
        capital=Coalesce(Sum('capital_principal_aplicado'), _money_zero()),
        interes=Coalesce(Sum('interes_aplicado'), _money_zero()),
        comision=Coalesce(Sum('comision_aplicada'), _money_zero()),
        iva=Coalesce(Sum('iva_aplicado'), _money_zero()),
    )
    rows = [
        ('Fecha generación', _excel_datetime(timezone.now())),
        ('Fecha desde', filtros.fecha_desde or 'Todo'),
        ('Fecha hasta', filtros.fecha_hasta or 'Todo'),
        ('Empresa', filtros.empresa.nombre if filtros.empresa else 'Todas'),
        ('Estado', filtros.estado or 'Todos'),
        ('Línea / producto', filtros.linea or 'Todas'),
        ('Ejecutivo / asesor', filtros.asesor.nombre if filtros.asesor else 'Todos'),
        ('Total solicitudes del período', solicitudes.count()),
        ('Total créditos desembolsados del período', desembolsos.count()),
        ('Monto desembolsado', desembolsos_totales['monto']),
        ('Pagos exitosos del período', pagos.count()),
        ('Total pagos del período', pagos_totales['monto']),
        ('Total recaudado', recaudo['total']),
        ('Capital recuperado', recaudo['capital']),
        ('Interés recuperado', recaudo['interes']),
        ('Comisión recuperada', recaudo['comision']),
        ('IVA recuperado', recaudo['iva']),
        ('Saldo total cartera - corte actual', cartera_actual['saldo']),
        ('Saldo capital pendiente - corte actual', cartera_actual['capital']),
        ('Semántica solicitudes', 'Filtradas por fecha de solicitud.'),
        ('Semántica créditos', 'Desembolsos filtrados por fecha de desembolso; saldos al corte actual.'),
        ('Semántica pagos', 'Pagos y recaudo filtrados por fecha de aplicación.'),
        ('Semántica cuotas', 'Cuotas filtradas por fecha de vencimiento.'),
    ]
    sheet = workbook.create_sheet('Resumen')
    _write_table(sheet, ['Concepto', 'Valor'], rows)
    for row in range(2, sheet.max_row + 1):
        label = sheet.cell(row=row, column=1).value
        cell = sheet.cell(row=row, column=2)
        if label == 'Fecha generación':
            cell.number_format = DATETIME_FORMAT
        elif label in {'Fecha desde', 'Fecha hasta'} and not isinstance(cell.value, str):
            cell.number_format = DATE_FORMAT
        elif label in {
            'Monto desembolsado',
            'Total pagos del período',
            'Total recaudado',
            'Capital recuperado',
            'Interés recuperado',
            'Comisión recuperada',
            'IVA recuperado',
            'Saldo total cartera - corte actual',
            'Saldo capital pendiente - corte actual',
        }:
            cell.number_format = MONEY_FORMAT


def _write_solicitudes(workbook, queryset):
    headers = [
        'Número solicitud / crédito',
        'Fecha solicitud',
        'Solicitante',
        'Documento enmascarado',
        'Empresa',
        'Línea / producto',
        'Monto solicitado',
        'Monto aprobado',
        'Plazo',
        'Estado actual',
        'Ejecutivo / asesor',
        'Fecha aprobación',
        'Fecha rechazo',
        'Fecha desembolso',
    ]
    rows = (
        (
            credito.numero_credito or f'SOL-{credito.pk}',
            _excel_datetime(credito.fecha_solicitud),
            credito.nombre_cliente,
            _mask_document(credito.cliente_documento),
            _company_name(credito),
            credito.get_linea_display(),
            credito.monto_solicitado,
            credito.monto_aprobado,
            credito.plazo or credito.plazo_solicitado,
            credito.get_estado_display(),
            _advisor_name(credito),
            _excel_datetime(credito.fecha_aprobacion_reporte),
            _excel_datetime(credito.fecha_rechazo_reporte),
            _excel_datetime(credito.fecha_desembolso),
        )
        for credito in queryset
    )
    _write_table(workbook.create_sheet('Solicitudes'), headers, rows)


def _write_creditos(workbook, queryset):
    headers = [
        'Número crédito',
        'Cliente',
        'Empresa',
        'Línea / producto',
        'Fecha solicitud',
        'Fecha desembolso',
        'Monto aprobado',
        'Plazo',
        'Tasa',
        'Estado actual',
        'Saldo total actual',
        'Capital pendiente actual',
        'Próxima fecha pago',
        'Ejecutivo / asesor',
    ]
    rows = (
        (
            credito.numero_credito,
            credito.nombre_cliente,
            _company_name(credito),
            credito.get_linea_display(),
            _excel_datetime(credito.fecha_solicitud),
            _excel_datetime(credito.fecha_desembolso),
            credito.monto_aprobado,
            credito.plazo,
            (credito.tasa_interes / Decimal('100')) if credito.tasa_interes is not None else None,
            credito.get_estado_display(),
            credito.saldo_pendiente,
            credito.capital_pendiente,
            credito.fecha_proximo_pago,
            _advisor_name(credito),
        )
        for credito in queryset
    )
    _write_table(workbook.create_sheet('Creditos'), headers, rows, percent_headers={'Tasa'})


def _write_cuotas(workbook, queryset):
    headers = [
        'Número crédito',
        'Empresa',
        'Cliente',
        'Número cuota',
        'Vencimiento',
        'Valor cuota',
        'Monto pagado',
        'Saldo pendiente cuota',
        'Pagada',
        'Estado operativo',
        'Días vencida',
        'Días para vencer',
    ]
    rows = []
    for cuota in queryset:
        estado, dias_vencida, dias_para_vencer = _installment_status(cuota)
        monto_pagado = cuota.monto_pagado or Decimal('0.00')
        rows.append((
            cuota.credito.numero_credito,
            _company_name(cuota.credito),
            cuota.credito.nombre_cliente,
            cuota.numero_cuota,
            cuota.fecha_vencimiento,
            cuota.valor_cuota,
            monto_pagado,
            max((cuota.valor_cuota or Decimal('0.00')) - monto_pagado, Decimal('0.00')),
            'Sí' if cuota.pagada else 'No',
            estado,
            dias_vencida,
            dias_para_vencer,
        ))
    _write_table(workbook.create_sheet('Cuotas'), headers, rows)


def _write_pagos(workbook, queryset):
    headers = [
        'Número crédito',
        'Cliente',
        'Empresa',
        'Fecha aplicación',
        'Valor total',
        'Capital',
        'Interés',
        'Comisión',
        'IVA',
        'Referencia / transacción',
        'Medio de pago',
        'Origen de pago',
        'Estado',
    ]
    rows = (
        (
            pago.credito.numero_credito,
            pago.credito.nombre_cliente,
            _company_name(pago.credito),
            _excel_datetime(pago.fecha_aplicacion),
            pago.monto,
            pago.capital_reporte,
            pago.interes_reporte,
            pago.comision_reporte,
            pago.iva_reporte,
            pago.referencia_pago,
            pago.get_metodo_pago_display(),
            pago.get_origen_registro_display(),
            pago.get_estado_display(),
        )
        for pago in queryset
    )
    _write_table(workbook.create_sheet('Pagos'), headers, rows)


def _write_secondary_sheets(workbook, base_creditos, detalles):
    cartera = base_creditos.filter(estado__in=ESTADOS_CARTERA)
    por_linea = cartera.values('linea').annotate(
        creditos=Count('pk', distinct=True),
        saldo=Coalesce(Sum('saldo_pendiente'), _money_zero()),
    ).order_by('linea')
    _write_table(
        workbook.create_sheet('Cartera por linea'),
        ['Línea', 'Créditos activos', 'Saldo total actual'],
        (
            (dict(Credito.LineaCredito.choices).get(item['linea'], item['linea']), item['creditos'], item['saldo'])
            for item in por_linea
        ),
        currency_headers={'Saldo total actual'},
    )

    total = base_creditos.count()
    por_estado = base_creditos.values('estado').annotate(cantidad=Count('pk', distinct=True)).order_by('estado')
    _write_table(
        workbook.create_sheet('Creditos por estado'),
        ['Estado actual', 'Cantidad', 'Porcentaje'],
        (
            (
                dict(Credito.EstadoCredito.choices).get(item['estado'], item['estado']),
                item['cantidad'],
                Decimal(item['cantidad']) / Decimal(total) if total else Decimal('0.00'),
            )
            for item in por_estado
        ),
        percent_headers={'Porcentaje'},
    )

    por_empresa = cartera.values('empresa_nombre').annotate(
        creditos=Count('pk', distinct=True),
        saldo=Coalesce(Sum('saldo_pendiente'), _money_zero()),
    ).order_by('-saldo', 'empresa_nombre')
    _write_table(
        workbook.create_sheet('Distribucion empresas'),
        ['Empresa', 'Créditos', 'Saldo total actual'],
        ((item['empresa_nombre'], item['creditos'], item['saldo']) for item in por_empresa),
        currency_headers={'Saldo total actual'},
    )

    por_credito = detalles.values(
        'credito__numero_credito',
        'credito__linea',
    ).annotate(
        total=Coalesce(Sum('monto_total_aplicado'), _money_zero()),
        capital=Coalesce(Sum('capital_principal_aplicado'), _money_zero()),
        interes=Coalesce(Sum('interes_aplicado'), _money_zero()),
        comision=Coalesce(Sum('comision_aplicada'), _money_zero()),
        iva=Coalesce(Sum('iva_aplicado'), _money_zero()),
    ).order_by('credito__numero_credito')
    _write_table(
        workbook.create_sheet('Recaudo contable'),
        [
            'Número crédito',
            'Línea / producto',
            'Total recaudado',
            'Capital recuperado',
            'Interés recuperado',
            'Comisión recuperada',
            'IVA recuperado',
        ],
        (
            (
                item['credito__numero_credito'],
                dict(Credito.LineaCredito.choices).get(item['credito__linea'], item['credito__linea']),
                item['total'],
                item['capital'],
                item['interes'],
                item['comision'],
                item['iva'],
            )
            for item in por_credito
        ),
    )

    _write_table(
        workbook.create_sheet('Detalle contable'),
        [
            'Número crédito',
            'Referencia / transacción',
            'Fecha aplicación',
            'Número cuota',
            'Secuencia',
            'Valor total',
            'Capital',
            'Interés',
            'Comisión',
            'IVA',
        ],
        (
            (
                detalle.credito.numero_credito,
                detalle.pago.referencia_pago,
                _excel_datetime(detalle.fecha_aplicacion),
                detalle.cuota.numero_cuota if detalle.cuota_id else None,
                detalle.secuencia_aplicacion,
                detalle.monto_total_aplicado,
                detalle.capital_principal_aplicado,
                detalle.interes_aplicado,
                detalle.comision_aplicada,
                detalle.iva_aplicado,
            )
            for detalle in detalles
        ),
    )


def _write_table(
    sheet,
    headers,
    rows,
    *,
    currency_headers=None,
    percent_headers=None,
):
    currency_headers = CURRENCY_HEADERS | set(currency_headers or ())
    percent_headers = set(percent_headers or ())
    sheet.append(headers)
    for row in rows:
        sheet.append([_safe_excel_value(value) for value in row])

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical='center')
    sheet.freeze_panes = 'A2'
    if sheet.max_column:
        sheet.auto_filter.ref = f'A1:{get_column_letter(sheet.max_column)}{max(sheet.max_row, 1)}'

    for column, header in enumerate(headers, start=1):
        number_format = None
        if header in currency_headers:
            number_format = MONEY_FORMAT
        elif header in percent_headers:
            number_format = PERCENT_FORMAT
        elif header in DATE_HEADERS:
            number_format = DATETIME_FORMAT if 'Fecha ' in header and header != 'Fecha desembolso' else DATE_FORMAT
        if number_format:
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row=row, column=column).number_format = number_format

        values = [sheet.cell(row=row, column=column).value for row in range(1, min(sheet.max_row, 80) + 1)]
        width = max((len(str(value or '')) for value in values), default=10)
        sheet.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 12), 38)


def _money_zero():
    return Value(
        Decimal('0.00'),
        output_field=DecimalField(max_digits=14, decimal_places=2),
    )


def _safe_excel_value(value):
    if isinstance(value, str) and value.startswith(('=', '+', '-', '@')):
        return f"'{value}"
    return value


def _excel_datetime(value):
    if value is None:
        return None
    if timezone.is_aware(value):
        return timezone.localtime(value).replace(tzinfo=None)
    return value


def _mask_document(value):
    value = ''.join(character for character in str(value or '') if character.isalnum())
    if not value:
        return ''
    return f"{'*' * max(len(value) - 4, 0)}{value[-4:]}"


def _company_for_credit(credito):
    if credito.linea == Credito.LineaCredito.LIBRANZA:
        detail = getattr(credito, 'detalle_libranza', None)
        return getattr(detail, 'empresa', None)
    if credito.linea == Credito.LineaCredito.ADELANTO_NOMINA:
        detail = getattr(credito, 'detalle_adelanto_nomina', None)
        vinculo = getattr(detail, 'vinculo_laboral', None)
        return getattr(vinculo, 'empresa', None)
    return None


def _company_name(credito):
    empresa = _company_for_credit(credito)
    return empresa.nombre if empresa else 'SIN EMPRESA'


def _advisor_name(credito):
    empresa = _company_for_credit(credito)
    asesor = getattr(empresa, 'asesor_comercial', None) if empresa else None
    return asesor.nombre if asesor else 'SIN EJECUTIVO'


def _installment_status(cuota):
    if cuota.pagada:
        return 'PAGADA', None, None
    hoy = timezone.localdate()
    diferencia = (cuota.fecha_vencimiento - hoy).days
    if diferencia < 0:
        return 'VENCIDA', abs(diferencia), None
    if diferencia == 0:
        return 'VENCE HOY', 0, 0
    if diferencia <= 15:
        return 'VENCE PRONTO', None, diferencia
    return 'AL DÍA', None, diferencia


def _report_range_suffix(filtros):
    if filtros.fecha_desde or filtros.fecha_hasta:
        start = filtros.fecha_desde.isoformat() if filtros.fecha_desde else 'inicio'
        end = filtros.fecha_hasta.isoformat() if filtros.fecha_hasta else 'hoy'
        return f'{start}_{end}'
    return f'todo_{timezone.localdate().isoformat()}'
